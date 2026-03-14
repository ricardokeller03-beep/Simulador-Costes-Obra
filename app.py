import streamlit as st
import pandas as pd
import openpyxl
import io
import copy

# ═══════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Simulador Costes Obra",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════════
# SECTION DEFINITIONS - row ranges and chapter names
# These define where each chapter starts/ends in the COSTOS sheet
# ═══════════════════════════════════════════════════════════════════
CHAPTER_DEFS = [
    {"key": "personal",     "label": "PERSONAL",              "start": 21, "end": 35,  "subtotal_row": 51, "driver_default": "tiempo"},
    {"key": "armadura",     "label": "ARMADURA",              "start": 53, "end": 66,  "subtotal_row": 67, "driver_default": "medicion"},
    {"key": "cemento",      "label": "CEMENTO",               "start": 69, "end": 73,  "subtotal_row": 74, "driver_default": "medicion"},
    {"key": "materiales",   "label": "OTROS MATERIALES",      "start": 76, "end": 90,  "subtotal_row": 91, "driver_default": "medicion"},
    {"key": "subcontrata",  "label": "SUBCONTRATA",           "start": 95, "end": 107, "subtotal_row": 108, "driver_default": "tiempo"},
    {"key": "maq_externa",  "label": "MAQUINARIA EXTERNA",    "start": 109, "end": 115, "subtotal_row": 116, "driver_default": "tiempo"},
    {"key": "otros_alq",    "label": "OTROS ALQUILERES",      "start": 125, "end": 133, "subtotal_row": 134, "driver_default": "tiempo"},
    {"key": "consumibles",  "label": "CONSUMIBLES / PARQUE",  "start": 137, "end": 152, "subtotal_row": 154, "driver_default": "medicion"},
    {"key": "gastos_var",   "label": "GASTOS VARIOS",         "start": 156, "end": 167, "subtotal_row": 168, "driver_default": "pct_prod"},
    {"key": "gasoil",       "label": "GASOIL",                "start": 172, "end": 173, "subtotal_row": 174, "driver_default": "pa"},
    {"key": "transportes",  "label": "TRANSPORTES",           "start": 176, "end": 177, "subtotal_row": 178, "driver_default": "pa"},
    {"key": "maq_interna",  "label": "MAQUINARIA INTERNA",    "start": 180, "end": 181, "subtotal_row": 182, "driver_default": "pa"},
]

PROD_DEF = {"start": 188, "end": 210, "subtotal_row": 211}

# Driver inference rules
def infer_driver(nombre, med, pu, chapter_default, base_dias_adj, base_ml, prod_total):
    """Infer the cost driver for a budget line based on its characteristics."""
    nombre_lower = nombre.lower() if nombre else ""
    
    # % sobre producción: medición ≈ producción total and PU < 0.1
    if med and prod_total and abs(med - prod_total) < 1000 and pu and pu < 0.1:
        return "pct_prod"
    
    # PA indicators: PU=1 and large med, or very small mediciones (1-5)
    if pu == 1 and med and med > 100:
        return "pa"
    if "pa " in nombre_lower or "partida" in nombre_lower:
        return "pa"
    
    # Time-based: medición is close to multiples of base_dias_adj
    if chapter_default == "tiempo" and med:
        ratio = med / base_dias_adj
        if abs(ratio - round(ratio)) < 0.3 and round(ratio) >= 1:
            return "tiempo"
    
    # If medición matches ML-based pattern
    if chapter_default == "medicion":
        return "medicion"
    
    return chapter_default


# ═══════════════════════════════════════════════════════════════════
# EXCEL PARSER
# ═══════════════════════════════════════════════════════════════════
@st.cache_data
def parse_excel(file_bytes):
    """Parse the COSTOS sheet from an .xlsm file and extract all budget data."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    
    # Find the costs sheet (COSTOS or COSTES or first sheet)
    sheet_name = None
    for name in wb.sheetnames:
        if name.upper() in ("COSTOS", "COSTES", "COSTS"):
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    
    ws = wb[sheet_name]
    
    # Extract header params
    params = {
        "nombre_obra": "",
        "expediente": "",
        "num_equipos": 2,
        "total_ml": 0,
        "rendimiento": 0,
        "dias_teoricos": 0,
    }
    
    # Read header rows
    for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=9, values_only=False):
        rn = row[0].row
        if rn == 3:
            params["nombre_obra"] = str(row[3].value or "").strip()
        if rn == 2:
            params["expediente"] = str(row[3].value or "").strip()
        if rn == 8:
            params["num_equipos"] = row[2].value or 2
        if rn == 9:
            try:
                params["total_ml"] = float(row[2].value or 0)
            except:
                params["total_ml"] = 0
        if rn == 10:
            try:
                params["rendimiento"] = float(row[2].value or 0)
            except:
                params["rendimiento"] = 0
        if rn == 11:
            try:
                params["dias_teoricos"] = float(row[2].value or 0)
            except:
                params["dias_teoricos"] = 0
    
    # Compute base dias ajustados (from personal section: max medicion / some integer)
    # We'll refine this after reading personal data
    
    # Read all rows into a dict for easy access
    all_rows = {}
    for row in ws.iter_rows(min_row=1, max_row=220, min_col=1, max_col=18, values_only=False):
        rn = None
        for c in row:
            if hasattr(c, 'row') and c.row:
                rn = c.row
                break
        if rn is None:
            continue
        all_rows[rn] = {
            "B": row[1].value,
            "C": row[2].value,
            "D": row[3].value,
            "E": row[4].value,
            "F": row[5].value,
            "G": row[6].value,
            "H": row[7].value,
            "I": row[8].value,
        }
    
    wb.close()
    
    # First pass: get production total for pct_prod detection
    prod_total = 0
    for rn in range(PROD_DEF["start"], PROD_DEF["end"] + 1):
        r = all_rows.get(rn, {})
        h = r.get("H")
        g = r.get("G")
        if h and g and isinstance(h, (int, float)) and isinstance(g, (int, float)):
            prod_total += h * g
    
    # Determine base_dias_ajustados from personal section
    # Look at the first personal line (usually Jefe Obra with 1 person = dias_ajustados)
    base_dias_adj = 24  # default
    for rn in range(21, 35):
        r = all_rows.get(rn, {})
        g = r.get("G")
        if g and isinstance(g, (int, float)) and g > 0:
            base_dias_adj = g  # First personal line = 1 person × dias
            break
    
    # Parse chapters
    chapters = []
    for cdef in CHAPTER_DEFS:
        items = []
        for rn in range(cdef["start"], cdef["end"] + 1):
            r = all_rows.get(rn, {})
            b = r.get("B")
            h = r.get("H")
            g = r.get("G")
            i_val = r.get("I")
            
            # Option 1: line has H and G data
            if h is not None and g is not None and isinstance(h, (int, float)) and isinstance(g, (int, float)):
                if h * g == 0:
                    continue
                nombre = str(b or "").strip()
                driver = infer_driver(
                    nombre, g, h, cdef["driver_default"],
                    base_dias_adj, params["total_ml"], prod_total
                )
                personas = None
                if cdef["key"] == "personal" and driver == "tiempo":
                    if base_dias_adj > 0:
                        personas = round(g / base_dias_adj)
                        if personas == 0:
                            personas = None
                items.append({
                    "row": rn,
                    "nombre": nombre if nombre and nombre.lower() not in ("0",) else f"Línea {rn}",
                    "pu": round(h, 4),
                    "med": round(g, 4),
                    "total": round(h * g, 2),
                    "driver": driver,
                    "personas": personas,
                    "ud": str(r.get("E") or ""),
                })
            # Option 2: line only has I value (subtotal-only, like Parque items or Maq Interna)
            elif i_val is not None and isinstance(i_val, (int, float)) and i_val != 0:
                nombre = str(b or "").strip()
                if nombre and nombre.lower() not in ("", "libre", "local", "ejemplo cem"):
                    # Check it's not a chapter subtotal we already track
                    is_subtotal = any(rn == cd["subtotal_row"] for cd in CHAPTER_DEFS)
                    if not is_subtotal:
                        items.append({
                            "row": rn,
                            "nombre": nombre,
                            "pu": 1,
                            "med": round(i_val, 2),
                            "total": round(i_val, 2),
                            "driver": "pa",
                            "personas": None,
                            "ud": str(r.get("E") or ""),
                        })
        
        # Get subtotal from Excel
        sr = all_rows.get(cdef["subtotal_row"], {})
        subtotal_excel = sr.get("I", 0)
        if not isinstance(subtotal_excel, (int, float)):
            subtotal_excel = sum(it["total"] for it in items)
        
        # If chapter has no items but has a subtotal, create a single PA item
        if not items and isinstance(subtotal_excel, (int, float)) and subtotal_excel > 0:
            sr_b = sr.get("B", cdef["label"])
            items.append({
                "row": cdef["subtotal_row"],
                "nombre": str(sr_b or cdef["label"]).strip(),
                "pu": 1,
                "med": round(subtotal_excel, 2),
                "total": round(subtotal_excel, 2),
                "driver": "pa",
                "personas": None,
                "ud": "€",
            })
        
        chapters.append({
            "key": cdef["key"],
            "label": cdef["label"],
            "items": items,
            "subtotal": round(subtotal_excel, 2),
            "driver_default": cdef["driver_default"],
        })
    
    # Parse production
    prod_items = []
    for rn in range(PROD_DEF["start"], PROD_DEF["end"] + 1):
        r = all_rows.get(rn, {})
        b = r.get("B")
        h = r.get("H")
        g = r.get("G")
        
        if h is None or g is None:
            continue
        if not isinstance(h, (int, float)) or not isinstance(g, (int, float)):
            continue
        if h == 0 and g == 0:
            continue
        
        nombre = str(b or "").strip()
        if not nombre or nombre.lower() in ("excesos",):
            if h * g == 0:
                continue
        
        prod_items.append({
            "row": rn,
            "nombre": nombre,
            "pu": round(h, 4),
            "med": round(g, 4),
            "total": round(h * g, 2),
        })
    
    return {
        "params": params,
        "chapters": chapters,
        "production": prod_items,
        "base_dias_adj": base_dias_adj,
        "prod_total": round(prod_total, 2),
    }


# ═══════════════════════════════════════════════════════════════════
# SIMULATION ENGINE
# ═══════════════════════════════════════════════════════════════════
def simulate(data, sim_params, sim_chapters, sim_production):
    """Recalculate all values based on simulation parameters."""
    base_ml = data["params"]["total_ml"]
    base_dias_adj = data["base_dias_adj"]
    
    rend_total = sim_params["rendimiento"] * sim_params["num_maquinas"]
    dias_teor = base_ml / rend_total if rend_total > 0 else 0
    # Use simulated ML for new dias
    new_rend_total = sim_params["rendimiento"] * sim_params["num_maquinas"]
    new_dias_teor = sim_params["total_ml"] / new_rend_total if new_rend_total > 0 else 0
    new_dias_adj = new_dias_teor + sim_params["dias_extra"]
    
    factor_dias = new_dias_adj / base_dias_adj if base_dias_adj > 0 else 1
    factor_ml = sim_params["total_ml"] / base_ml if base_ml > 0 else 1
    
    # Production total
    total_prod = sum(p["pu"] * p["med"] for p in sim_production)
    
    # Recalculate chapters
    result_chapters = []
    total_coste = 0
    
    for ci, chapter in enumerate(sim_chapters):
        items = []
        for item in chapter["items"]:
            med_calc = item["med"]
            orig_item = data["chapters"][ci]["items"]
            orig_med = item.get("med_base", item["med"])
            
            driver = item["driver"]
            if driver == "tiempo":
                if item.get("personas") and item["personas"] > 0:
                    med_calc = item["personas"] * new_dias_adj
                else:
                    med_calc = orig_med * factor_dias
            elif driver == "medicion":
                med_calc = orig_med * factor_ml
            elif driver == "pct_prod":
                med_calc = total_prod
            # pa stays as is
            
            total = item["pu"] * med_calc
            items.append({**item, "med_calc": round(med_calc, 2), "total_sim": round(total, 2)})
        
        chapter_total = sum(it["total_sim"] for it in items)
        total_coste += chapter_total
        result_chapters.append({**chapter, "items": items, "total_sim": round(chapter_total, 2)})
    
    margen = total_prod - total_coste
    pct_mb = margen / total_prod if total_prod > 0 else 0
    
    return {
        "dias_teor": round(new_dias_teor, 2),
        "dias_adj": round(new_dias_adj, 2),
        "rend_total": round(new_rend_total, 2),
        "factor_dias": round(factor_dias, 4),
        "factor_ml": round(factor_ml, 4),
        "total_prod": round(total_prod, 2),
        "total_coste": round(total_coste, 2),
        "margen": round(margen, 2),
        "pct_mb": round(pct_mb, 6),
        "chapters": result_chapters,
    }


# ═══════════════════════════════════════════════════════════════════
# FORMATTING
# ═══════════════════════════════════════════════════════════════════
def fmt_eur(n):
    if n is None: return "—"
    return f"{n:,.0f} €".replace(",", ".")

def fmt_pct(n):
    return f"{n*100:.2f}%"

def desv_color(val):
    if abs(val) < 1: return "gray"
    return "red" if val > 0 else "green"

def desv_str(val):
    if abs(val) < 1: return "—"
    sign = "+" if val > 0 else ""
    return f"{sign}{val:,.0f} €".replace(",", ".")


# ═══════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════
def main():
    # Custom CSS
    st.markdown("""
    <style>
    .block-container { padding: 1rem 2rem; }
    .stMetric { background: #0f172a; border-radius: 8px; padding: 12px; }
    div[data-testid="stMetricValue"] { font-size: 1.8rem; }
    .section-header {
        background: #1e293b; border-radius: 6px; padding: 8px 12px;
        margin: 8px 0 4px 0; display: flex; justify-content: space-between;
    }
    table { font-size: 0.82rem !important; }
    th { background: #1e293b !important; font-size: 0.75rem !important; }
    </style>
    """, unsafe_allow_html=True)
    
    st.title("🏗️ Simulador de Presupuesto de Obra")
    
    # File upload
    uploaded = st.file_uploader(
        "Sube el fichero .xlsm de costes",
        type=["xlsm", "xlsx"],
        help="El fichero debe tener una hoja COSTOS/COSTES con la estructura estándar"
    )
    
    if not uploaded:
        st.info("👆 Sube un fichero Excel de costes para comenzar. Debe tener una hoja COSTOS con la estructura estándar de presupuesto.")
        st.stop()
    
    # Parse
    data = parse_excel(uploaded.read())
    p = data["params"]
    
    st.caption(f"**{p['nombre_obra']}** · Exp. {p['expediente']} · {p['total_ml']} ML · {p['num_equipos']} equipos")
    
    # ─── SIDEBAR: Parameters ───────────────────────────────────
    st.sidebar.header("⚙️ Parámetros de Simulación")
    
    sim_ml = st.sidebar.number_input("ML Totales", value=float(p["total_ml"]), step=10.0, format="%.1f")
    sim_rend = st.sidebar.number_input("Rendimiento ML/día/máquina", value=float(p["rendimiento"]), step=1.0, format="%.2f")
    sim_maq = st.sidebar.number_input("Nº Máquinas", value=int(p["num_equipos"]), min_value=1, step=1)
    
    base_dias_extra = round(data["base_dias_adj"] - (p["total_ml"] / (p["rendimiento"] * p["num_equipos"])) if p["rendimiento"] * p["num_equipos"] > 0 else 4, 1)
    sim_dias_extra = st.sidebar.number_input("Días extra (montaje/avería)", value=float(base_dias_extra), step=0.5, format="%.1f")
    
    sim_params = {
        "total_ml": sim_ml,
        "rendimiento": sim_rend,
        "num_maquinas": sim_maq,
        "dias_extra": sim_dias_extra,
    }
    
    rend_total = sim_rend * sim_maq
    dias_teor = sim_ml / rend_total if rend_total > 0 else 0
    dias_adj = dias_teor + sim_dias_extra
    
    st.sidebar.markdown(f"""
    ---
    **Rendimiento total:** {rend_total:.1f} ML/día  
    **Días teóricos:** {dias_teor:.1f}  
    **Días ajustados:** {dias_adj:.1f}  
    """)
    
    # ─── Initialize session state for editable values ──────────
    if "sim_chapters" not in st.session_state or st.sidebar.button("🔄 Reset a valores originales"):
        sim_chapters = []
        for ch in data["chapters"]:
            items = []
            for it in ch["items"]:
                items.append({**it, "med_base": it["med"]})
            sim_chapters.append({**ch, "items": items})
        st.session_state.sim_chapters = sim_chapters
        st.session_state.sim_production = copy.deepcopy(data["production"])
    
    # ─── Run simulation ────────────────────────────────────────
    result = simulate(data, sim_params, st.session_state.sim_chapters, st.session_state.sim_production)
    
    # Original totals
    orig_coste = sum(ch["subtotal"] for ch in data["chapters"])
    orig_prod = data["prod_total"]
    orig_mb = orig_prod - orig_coste
    orig_pct = orig_mb / orig_prod if orig_prod > 0 else 0
    
    # ─── KPIs ──────────────────────────────────────────────────
    k1, k2, k3, k4 = st.columns(4)
    
    with k1:
        delta_mb = result["pct_mb"] - orig_pct
        st.metric(
            "% Margen Bruto",
            fmt_pct(result["pct_mb"]),
            f"{delta_mb*100:+.2f}pp (orig: {fmt_pct(orig_pct)})",
            delta_color="normal",
        )
    with k2:
        st.metric(
            "Margen Bruto",
            fmt_eur(result["margen"]),
            f"{desv_str(result['margen'] - orig_mb)} vs orig",
            delta_color="normal",
        )
    with k3:
        st.metric(
            "Coste Directo",
            fmt_eur(result["total_coste"]),
            f"{desv_str(result['total_coste'] - orig_coste)} vs orig",
            delta_color="inverse",
        )
    with k4:
        st.metric(
            "Producción",
            fmt_eur(result["total_prod"]),
            f"{sim_maq} máq · {rend_total:.0f} ML/d · {dias_adj:.1f} días",
        )
    
    # ─── TABS ──────────────────────────────────────────────────
    tab_costes, tab_prod, tab_resumen = st.tabs(["📊 Costes por Capítulo", "💰 Producción", "📋 Resumen Comparativo"])
    
    # ═══ TAB: COSTES ═══
    with tab_costes:
        st.caption("Izquierda = Original (fijo) · Derecha = Simulado (editable) · Las mediciones de TIEMPO y MEDICIÓN se recalculan automáticamente")
        
        for ci, chapter in enumerate(result["chapters"]):
            orig_ch = data["chapters"][ci]
            desv_ch = chapter["total_sim"] - orig_ch["subtotal"]
            desv_pct = desv_ch / orig_ch["subtotal"] * 100 if orig_ch["subtotal"] > 0 else 0
            
            color = "🔵" if orig_ch["driver_default"] == "tiempo" else "🟢" if orig_ch["driver_default"] == "medicion" else "🟡" if orig_ch["driver_default"] == "pa" else "🩷"
            
            with st.expander(
                f"{color} **{chapter['label']}** — Orig: {fmt_eur(orig_ch['subtotal'])} → Sim: {fmt_eur(chapter['total_sim'])} ({desv_pct:+.1f}%)",
                expanded=False
            ):
                # Build comparison dataframe
                rows = []
                for ii, item in enumerate(chapter["items"]):
                    orig_it = orig_ch["items"][ii] if ii < len(orig_ch["items"]) else None
                    orig_total = orig_it["total"] if orig_it else 0
                    desv = item["total_sim"] - orig_total
                    
                    rows.append({
                        "Concepto": item["nombre"],
                        "Driver": item["driver"].upper(),
                        "PU orig": f"{orig_it['pu']:.2f}" if orig_it else "—",
                        "Med orig": f"{orig_it['med']:.1f}" if orig_it else "—",
                        "Total orig": fmt_eur(orig_total),
                        "│": "│",
                        "PU sim": item["pu"],
                        "Med sim": item["med_calc"],
                        "Total sim": fmt_eur(item["total_sim"]),
                        "Desviación": desv_str(desv),
                    })
                
                # Display as dataframe
                df_display = pd.DataFrame(rows)
                st.dataframe(df_display, use_container_width=True, hide_index=True)
                
                # Editable section
                st.markdown("**✏️ Editar valores simulados:**")
                cols_edit = st.columns([2, 1, 1, 1])
                
                for ii, item in enumerate(chapter["items"]):
                    with cols_edit[0]:
                        st.text(item["nombre"])
                    with cols_edit[1]:
                        new_pu = st.number_input(
                            f"PU_{ci}_{ii}", value=float(item["pu"]),
                            step=0.1, format="%.4f",
                            label_visibility="collapsed",
                            key=f"pu_{ci}_{ii}"
                        )
                        if new_pu != item["pu"]:
                            st.session_state.sim_chapters[ci]["items"][ii]["pu"] = new_pu
                            st.rerun()
                    
                    with cols_edit[2]:
                        if item["driver"] == "pa":
                            new_med = st.number_input(
                                f"Med_{ci}_{ii}", value=float(item["med"]),
                                step=1.0, format="%.1f",
                                label_visibility="collapsed",
                                key=f"med_{ci}_{ii}"
                            )
                            if new_med != item["med"]:
                                st.session_state.sim_chapters[ci]["items"][ii]["med"] = new_med
                                st.session_state.sim_chapters[ci]["items"][ii]["med_base"] = new_med
                                st.rerun()
                        else:
                            st.text(f"{item['med_calc']:.1f} (auto)")
                    
                    with cols_edit[3]:
                        if item.get("personas") is not None and chapter["key"] == "personal":
                            new_pers = st.number_input(
                                f"Pers_{ci}_{ii}", value=int(item["personas"]),
                                min_value=0, step=1,
                                label_visibility="collapsed",
                                key=f"pers_{ci}_{ii}"
                            )
                            if new_pers != item["personas"]:
                                st.session_state.sim_chapters[ci]["items"][ii]["personas"] = new_pers
                                st.rerun()
                
                # Subtotals
                st.markdown(f"""
                | | Original | Simulado | Desviación |
                |---|---:|---:|---:|
                | **TOTAL {chapter['label']}** | **{fmt_eur(orig_ch['subtotal'])}** | **{fmt_eur(chapter['total_sim'])}** | **{desv_str(desv_ch)}** |
                """)
        
        # TOTAL COSTO DIRECTO
        st.markdown("---")
        tc1, tc2, tc3 = st.columns(3)
        with tc1:
            st.metric("TOTAL COSTO DIRECTO - Original", fmt_eur(orig_coste))
        with tc2:
            st.metric("TOTAL COSTO DIRECTO - Simulado", fmt_eur(result["total_coste"]))
        with tc3:
            desv_total = result["total_coste"] - orig_coste
            st.metric("Desviación", desv_str(desv_total), f"{desv_total/orig_coste*100:+.1f}%" if orig_coste > 0 else "")
    
    # ═══ TAB: PRODUCCIÓN ═══
    with tab_prod:
        st.caption("Izquierda = Original (fijo) · Derecha = Simulado (editable)")
        
        rows = []
        for ii, prod in enumerate(st.session_state.sim_production):
            orig_p = data["production"][ii] if ii < len(data["production"]) else None
            orig_total = orig_p["total"] if orig_p else 0
            sim_total = prod["pu"] * prod["med"]
            desv = sim_total - orig_total
            
            rows.append({
                "Concepto": prod["nombre"],
                "PU orig": f"{orig_p['pu']:.2f}" if orig_p else "—",
                "Med orig": f"{orig_p['med']:.1f}" if orig_p else "—",
                "Total orig": fmt_eur(orig_total),
                "│": "│",
                "PU sim": f"{prod['pu']:.2f}",
                "Med sim": f"{prod['med']:.1f}",
                "Total sim": fmt_eur(sim_total),
                "Desviación": desv_str(desv),
            })
        
        df_prod = pd.DataFrame(rows)
        st.dataframe(df_prod, use_container_width=True, hide_index=True)
        
        st.markdown("**✏️ Editar producción:**")
        for ii, prod in enumerate(st.session_state.sim_production):
            c1, c2, c3 = st.columns([3, 1, 1])
            with c1:
                st.text(prod["nombre"])
            with c2:
                new_pu = st.number_input(
                    f"ProdPU_{ii}", value=float(prod["pu"]),
                    step=0.1, format="%.2f",
                    label_visibility="collapsed",
                    key=f"prod_pu_{ii}"
                )
                if new_pu != prod["pu"]:
                    st.session_state.sim_production[ii]["pu"] = new_pu
                    st.rerun()
            with c3:
                new_med = st.number_input(
                    f"ProdMed_{ii}", value=float(prod["med"]),
                    step=1.0, format="%.1f",
                    label_visibility="collapsed",
                    key=f"prod_med_{ii}"
                )
                if new_med != prod["med"]:
                    st.session_state.sim_production[ii]["med"] = new_med
                    st.rerun()
        
        st.markdown("---")
        p1, p2, p3 = st.columns(3)
        with p1:
            st.metric("Producción", fmt_eur(result["total_prod"]))
        with p2:
            st.metric("Coste Directo", fmt_eur(result["total_coste"]))
        with p3:
            st.metric("Margen Bruto", fmt_eur(result["margen"]), fmt_pct(result["pct_mb"]))
    
    # ═══ TAB: RESUMEN ═══
    with tab_resumen:
        summary_rows = []
        for ci, chapter in enumerate(result["chapters"]):
            orig_ch = data["chapters"][ci]
            desv = chapter["total_sim"] - orig_ch["subtotal"]
            desv_pct = desv / orig_ch["subtotal"] * 100 if orig_ch["subtotal"] > 0 else 0
            summary_rows.append({
                "Capítulo": chapter["label"],
                "Original": fmt_eur(orig_ch["subtotal"]),
                "Simulado": fmt_eur(chapter["total_sim"]),
                "Desviación": desv_str(desv),
                "% Desv.": f"{desv_pct:+.1f}%",
            })
        
        df_summary = pd.DataFrame(summary_rows)
        st.dataframe(df_summary, use_container_width=True, hide_index=True)
        
        st.markdown("---")
        
        # Final summary table
        summary_data = {
            "": ["TOTAL COSTO DIRECTO", "PRODUCCIÓN", "MARGEN BRUTO", "% MB"],
            "Original": [fmt_eur(orig_coste), fmt_eur(orig_prod), fmt_eur(orig_mb), fmt_pct(orig_pct)],
            "Simulado": [fmt_eur(result["total_coste"]), fmt_eur(result["total_prod"]), fmt_eur(result["margen"]), fmt_pct(result["pct_mb"])],
            "Desviación": [
                desv_str(result["total_coste"] - orig_coste),
                desv_str(result["total_prod"] - orig_prod),
                desv_str(result["margen"] - orig_mb),
                f"{(result['pct_mb'] - orig_pct)*100:+.2f}pp",
            ],
        }
        st.table(pd.DataFrame(summary_data))


if __name__ == "__main__":
    main()
